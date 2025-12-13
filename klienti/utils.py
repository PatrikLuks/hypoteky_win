import csv
import datetime
import logging
from typing import IO

from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string

import openpyxl

from .models import Klient, NotifikaceLog

# Logger pro import operace
logger = logging.getLogger("klienti.import")


def odeslat_notifikaci_email(
    prijemce,
    predmet,
    zprava,
    html_zprava=None,
    context=None,
    template_name=None,
    typ="deadline",
    klient=None,
):
    """
    Odeslání e-mailové notifikace s logováním do NotifikaceLog.
    """
    html_message = html_zprava
    if template_name and context:
        html_message = render_to_string(template_name, context)
    uspesne = True
    try:
        send_mail(
            subject=predmet,
            message=zprava,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[prijemce],
            html_message=html_message,
            fail_silently=False,
        )
    except Exception:
        uspesne = False
    NotifikaceLog.objects.create(
        prijemce=prijemce, typ=typ, klient=klient, obsah=zprava, uspesne=uspesne
    )


def import_klienti_from_csv(file: IO, encoding: str = "utf-8") -> int:
    """
    Načte klienty z CSV souboru a uloží je do DB.
    Očekává hlavičku: jmeno,datum,vyber_banky,navrh_financovani_castka,duvod_zamitnuti,co_financuje,navrh_financovani,navrh_financovani_procento,cena
    Vrací počet úspěšně importovaných klientů (nově vytvořených nebo aktualizovaných).
    Robustně ošetřuje poškozené nebo nevalidní CSV soubory – nikdy nespadne, pouze zaloguje chybu a přeskočí řádek.
    """
    try:
        reader = csv.DictReader(
            (
                line.decode(encoding) if isinstance(line, bytes) else line
                for line in file
            )
        )
    except Exception as e:
        logger.error(f"[IMPORT] Chyba při čtení CSV: {e}")
        return 0
    count = 0
    imported_keys = set()  # Sada (jmeno, datum) pro detekci duplicit v rámci dávky
    try:
        for idx, row in enumerate(reader):
            try:
                logger.debug(f"[IMPORT] Zpracovávám řádek {idx + 1}: {row}")
                if not row.get("jmeno"):
                    logger.warning(f"[IMPORT] Řádek {idx + 1} přeskočen - chybí jméno")
                    continue  # povinné pole
                if not row.get("vyber_banky"):
                    logger.warning(f"[IMPORT] Řádek {idx + 1} přeskočen - chybí vyber_banky")
                    continue  # povinné pole
                # --- Validace povinného pole datum ---
                jmeno = row["jmeno"].strip() if row.get("jmeno") else ""
                datum_str = row.get("datum", "").strip()
                try:
                    datum = datetime.datetime.strptime(datum_str, "%Y-%m-%d").date()
                except Exception:
                    logger.warning(f"[IMPORT] Chyba při převodu datumu: {datum_str}")
                    logger.warning(
                        f"[IMPORT] Řádek {idx + 1} přeskočen - chybí nebo je nevalidní datum"
                    )
                    continue
                key = (jmeno, datum)
                if key in imported_keys:
                    logger.warning(
                        f"[IMPORT] Řádek {idx + 1} přeskočen - duplicitní klient v dávce (jméno+datum)"
                    )
                    continue
                # Použití jmeno_index pro filtrování (jmeno je šifrované pole)
                if Klient.objects.filter(jmeno_index=jmeno, datum=datum).exists():
                    logger.info(
                        f"[IMPORT] Řádek {idx + 1} přeskočen - duplicitní klient v DB (jméno+datum)"
                    )
                    continue
                imported_keys.add(key)
                castka = None
                if row.get("navrh_financovani_castka"):
                    try:
                        castka = int(row["navrh_financovani_castka"])
                    except Exception as e:
                        logger.warning(
                            f"[IMPORT] Chyba při převodu částky: {row['navrh_financovani_castka']} ({e})"
                        )
                cena = None
                if row.get("cena"):
                    try:
                        cena = float(row["cena"])
                    except Exception as e:
                        logger.warning(f"[IMPORT] Chyba při převodu ceny: {row['cena']} ({e})")
                procento = None
                if row.get("navrh_financovani_procento"):
                    try:
                        procento = float(row["navrh_financovani_procento"])
                    except Exception as e:
                        logger.warning(
                            f"[IMPORT] Chyba při převodu procenta: {row['navrh_financovani_procento']} ({e})"
                        )
                klient = Klient(
                    jmeno=jmeno,
                    datum=datum,
                    vyber_banky=row.get("vyber_banky"),
                    navrh_financovani_castka=castka,
                    duvod_zamitnuti=row.get("duvod_zamitnuti") or None,
                    co_financuje=row.get("co_financuje") or "",
                    navrh_financovani=row.get("navrh_financovani") or "",
                    navrh_financovani_procento=procento,
                    cena=cena,
                )
                klient.save()
                logger.info(f"[IMPORT] Klient vytvořen: {klient.jmeno}")
                count += 1
            except Exception as e:
                logger.error(f"[IMPORT] Chyba při zpracování řádku {idx + 1}: {e}")
                continue
    except Exception as e:
        logger.error(f"[IMPORT] Chyba při čtení CSV v průběhu zpracování: {e}")
        return count
    logger.info(f"[IMPORT] Celkem importováno/aktualizováno klientů: {count}")
    return count


def import_klienti_from_xlsx(file: IO) -> int:
    """
    Načte klienty z XLSX souboru a uloží je do DB.
    Očekává hlavičku: jmeno,datum,vyber_banky,navrh_financovani_castka,duvod_zamitnuti,co_financuje,navrh_financovani,navrh_financovani_procento,cena
    Vrací počet úspěšně importovaných klientů (nově vytvořených nebo aktualizovaných).
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    count = 0
    imported_keys = set()  # Sada (jmeno, datum) pro detekci duplicit v rámci dávky
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {
            headers[i]: (cell.value if cell.value is not None else "")
            for i, cell in enumerate(row)
        }
        logger.debug(f"[IMPORT XLSX] Zpracovávám řádek {idx}: {row_data}")
        if not row_data.get("jmeno"):
            logger.warning(f"[IMPORT XLSX] Řádek {idx} přeskočen - chybí jméno")
            continue
        jmeno = row_data["jmeno"].strip()
        # Validace délky jména podle modelu (max_length=100)
        if len(jmeno) > 100:
            logger.warning(
                f"[IMPORT XLSX] VAROVÁNÍ: Jméno je delší než 100 znaků, bude zkráceno. Původní: {jmeno}"
            )
            jmeno = jmeno[:100]
        datum = None
        if row_data.get("datum"):
            try:
                datum = row_data["datum"]
                if isinstance(datum, str):
                    datum = datetime.datetime.strptime(datum, "%Y-%m-%d").date()
                elif isinstance(datum, datetime.datetime):
                    datum = datum.date()
            except Exception as e:
                logger.warning(
                    f"[IMPORT XLSX] Chyba při převodu datumu: {row_data['datum']} ({e})"
                )
        if not datum:
            logger.warning(
                f"[IMPORT XLSX] Řádek {idx} přeskočen - chybí nebo je nevalidní datum"
            )
            continue
        key = (jmeno, datum)
        if key in imported_keys:
            logger.warning(
                f"[IMPORT XLSX] Řádek {idx} přeskočen - duplicitní klient v dávce (jméno+datum)"
            )
            continue
        # Použití jmeno_index pro filtrování (jmeno je šifrované pole)
        if Klient.objects.filter(jmeno_index=jmeno, datum=datum).exists():
            logger.info(
                f"[IMPORT XLSX] Řádek {idx} přeskočen - duplicitní klient v DB (jméno+datum)"
            )
            continue
        imported_keys.add(key)
        castka = None
        if row_data.get("navrh_financovani_castka"):
            try:
                castka = int(row_data["navrh_financovani_castka"])
            except Exception as e:
                logger.warning(
                    f"[IMPORT XLSX] Chyba při převodu částky: {row_data['navrh_financovani_castka']} ({e})"
                )
        # Oprava: správně rozpoznat nulu i prázdné hodnoty pro pole cena
        cena = None
        if row_data.get("cena") is not None and row_data.get("cena") != "":
            try:
                cena = float(row_data["cena"])
            except Exception as e:
                logger.warning(f"[IMPORT XLSX] Chyba při převodu ceny: {row_data['cena']} ({e})")
        procento = None
        if row_data.get("navrh_financovani_procento"):
            try:
                procento = float(row_data["navrh_financovani_procento"])
            except Exception as e:
                logger.warning(
                    f"[IMPORT XLSX] Chyba při převodu procenta: {row_data['navrh_financovani_procento']} ({e})"
                )
        # Použití jmeno_index pro get_or_create (jmeno je šifrované pole)
        klient, created = Klient.objects.get_or_create(
            jmeno_index=jmeno,
            defaults={
                "jmeno": jmeno,
                "datum": datum,
                "vyber_banky": row_data.get("vyber_banky"),
                "navrh_financovani_castka": castka,
                "duvod_zamitnuti": row_data.get("duvod_zamitnuti") or None,
                "co_financuje": row_data.get("co_financuje") or "",
                "navrh_financovani": row_data.get("navrh_financovani") or "",
                "navrh_financovani_procento": procento,
                "cena": cena if cena is not None else 0,
            },
        )
        if not created:
            klient.datum = datum or klient.datum
            klient.vyber_banky = row_data.get("vyber_banky") or klient.vyber_banky
            klient.navrh_financovani_castka = (
                castka if castka is not None else klient.navrh_financovani_castka
            )
            klient.duvod_zamitnuti = (
                row_data.get("duvod_zamitnuti") or klient.duvod_zamitnuti
            )
            klient.co_financuje = row_data.get("co_financuje") or klient.co_financuje
            klient.navrh_financovani = (
                row_data.get("navrh_financovani") or klient.navrh_financovani
            )
            klient.navrh_financovani_procento = (
                procento if procento is not None else klient.navrh_financovani_procento
            )
            klient.cena = cena if cena is not None else klient.cena
            klient.save()
            logger.info(f"[IMPORT XLSX] Klient aktualizován: {klient.jmeno}")
        else:
            logger.info(f"[IMPORT XLSX] Klient vytvořen: {klient.jmeno}")
        count += 1
    logger.info(f"[IMPORT XLSX] Celkem importováno/aktualizováno klientů: {count}")
    return count
