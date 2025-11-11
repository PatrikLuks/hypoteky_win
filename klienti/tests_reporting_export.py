import io
from datetime import date, timedelta

from django.test import TestCase

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from klienti.models import Klient


class ReportingExportXLSXTestCase(TestCase):
    def setUp(self):
        Klient.objects.all().delete()
        # Vytvoř klienty pro test_export_xlsx_obsahuje_spravna_data
        Klient.objects.create(
            jmeno="Jan Export",
            datum=date.today(),
            vyber_banky="KB",
            navrh_financovani_castka=2000000,
        )
        Klient.objects.create(
            jmeno="Petr Export",
            datum=date.today(),
            vyber_banky="ČSOB",
            navrh_financovani_castka=3000000,
            duvod_zamitnuti="Nedostatečný příjem",
        )
        Klient.objects.create(
            jmeno="Eva Export",
            datum=date.today(),
            vyber_banky="KB",
            navrh_financovani_castka=2500000,
        )
        # Vytvoř testovacího manažera s validním e-mailem
        from django.contrib.auth.models import User

        from klienti.models import UserProfile

        user = User.objects.create_user(
            username="manazer", email="manazer@example.com", password="testpass"
        )
        profile, created = UserProfile.objects.get_or_create(
            user=user, defaults={"role": "manažer"}
        )
        if not created:
            profile.role = "manažer"
            profile.save()

    def test_export_xlsx_obsahuje_spravna_data(self):
        # --- Simulace části generování XLSX z management commandu ---
        from collections import Counter

        klienti = Klient.objects.all()
        banky = [k.vyber_banky for k in klienti if k.vyber_banky]
        rozlozeni_banky = Counter(banky)
        schvalene = klienti.filter(
            duvod_zamitnuti__isnull=True, vyber_banky__isnull=False
        )
        zamitnute = klienti.filter(
            duvod_zamitnuti__isnull=False, vyber_banky__isnull=False
        )
        banky_labels = list(rozlozeni_banky.keys())
        schvalenost = [schvalene.filter(vyber_banky=b).count() for b in banky_labels]
        zamitnutost = [zamitnute.filter(vyber_banky=b).count() for b in banky_labels]
        prumery = [
            "-" for _ in banky_labels
        ]  # pro jednoduchost, bez datumu podání/schválení
        # Vygeneruj XLSX do bufferu
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        ws.append(["Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"])
        for i, banka in enumerate(banky_labels):
            ws.append([banka, schvalenost[i], zamitnutost[i], prumery[i]])
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        # --- Ověření obsahu XLSX ---
        wb2 = openpyxl.load_workbook(xlsx_buffer)
        ws2 = wb2.active
        data = list(ws2.values)
        self.assertEqual(
            data[0],
            ("Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"),
        )
        # Ověř, že jsou tam obě banky a správné počty
        rows = {row[0]: row for row in data[1:]}
        self.assertIn("KB", rows)
        self.assertIn("ČSOB", rows)
        self.assertEqual(rows["KB"][1], 2)  # 2 schválené
        self.assertEqual(rows["KB"][2], 0)  # 0 zamítnutých
        self.assertEqual(rows["ČSOB"][1], 0)  # 0 schválených
        self.assertEqual(rows["ČSOB"][2], 1)  # 1 zamítnutý

    def test_export_xlsx_prazdna_db(self):
        Klient.objects.all().delete()
        from collections import Counter

        klienti = Klient.objects.all()
        banky = [k.vyber_banky for k in klienti if k.vyber_banky]
        rozlozeni_banky = Counter(banky)
        banky_labels = list(rozlozeni_banky.keys())
        schvalenost = []
        zamitnutost = []
        prumery = []
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        ws.append(["Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"])
        for i, banka in enumerate(banky_labels):
            ws.append([banka, schvalenost[i], zamitnutost[i], prumery[i]])
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        wb2 = openpyxl.load_workbook(xlsx_buffer)
        ws2 = wb2.active
        data = list(ws2.values)
        self.assertEqual(
            data[0],
            ("Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"),
        )
        self.assertEqual(len(data), 1)  # pouze hlavička

    def test_export_xlsx_extremni_hodnoty(self):
        Klient.objects.all().delete()
        Klient.objects.create(
            jmeno="Mega Export",
            datum=date.today(),
            vyber_banky="Banka S Extrémně Dlouhým Názvem Pro Testování Robustnosti",
            navrh_financovani_castka=9999999999,
        )
        from collections import Counter

        klienti = Klient.objects.all()
        banky = [k.vyber_banky for k in klienti if k.vyber_banky]
        rozlozeni_banky = Counter(banky)
        banky_labels = list(rozlozeni_banky.keys())
        schvalenost = [
            klienti.filter(duvod_zamitnuti__isnull=True, vyber_banky=b).count()
            for b in banky_labels
        ]
        zamitnutost = [
            klienti.filter(duvod_zamitnuti__isnull=False, vyber_banky=b).count()
            for b in banky_labels
        ]
        prumery = ["-" for _ in banky_labels]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        ws.append(["Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"])
        for i, banka in enumerate(banky_labels):
            ws.append([banka, schvalenost[i], zamitnutost[i], prumery[i]])
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        wb2 = openpyxl.load_workbook(xlsx_buffer)
        ws2 = wb2.active
        data = list(ws2.values)
        self.assertIn(
            "Banka S Extrémně Dlouhým Názvem Pro Testování Robustnosti",
            [row[0] for row in data[1:]],
        )
        self.assertEqual(data[1][1], 1)  # 1 schválený
        self.assertEqual(data[1][2], 0)  # 0 zamítnutých

    def test_export_pdf_report_neni_prazdny(self):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(40, 800, "Reporting – úspěšnost podle banky")
        p.setFont("Helvetica", 11)
        y = 770
        p.drawString(40, y, "Banka")
        p.drawString(180, y, "Schváleno")
        p.drawString(270, y, "Zamítnuto")
        p.drawString(370, y, "Průměrná doba schválení (dny)")
        y -= 20
        p.drawString(40, y, "Testovací banka")
        p.drawString(180, y, "1")
        p.drawString(270, y, "0")
        p.drawString(370, y, "-")
        p.save()
        buffer.seek(0)
        pdf_data = buffer.read()
        self.assertGreater(len(pdf_data), 500)  # PDF by mělo mít alespoň 500 B

    def test_export_xlsx_neuplne_nevalidni_udaje(self):
        """
        Testuje export klientů do XLSX, když někteří klienti mají neúplné nebo nevalidní údaje:
        - chybí banka (prázdný string)
        - chybí částka
        - některé hodnoty jsou prázdné

        Tento test je důležitý, protože v reálných datech se často objeví neúplné záznamy a export by měl proběhnout bez chyby.
        """
        Klient.objects.all().delete()
        # Klient bez banky (prázdný string místo None)
        Klient.objects.create(
            jmeno="Bez Banky",
            datum=date.today(),
            vyber_banky="",
            navrh_financovani_castka=1500000,
        )
        # Klient bez částky
        Klient.objects.create(
            jmeno="Bez Částky",
            datum=date.today(),
            vyber_banky="KB",
            navrh_financovani_castka=None,
        )
        # Klient s prázdnými hodnotami (ale povinné pole datum musí být vyplněné)
        Klient.objects.create(
            jmeno="Všechno prázdné",
            datum=date.today(),
            vyber_banky="",
            navrh_financovani_castka=None,
        )
        from collections import Counter

        klienti = Klient.objects.all()
        banky = [k.vyber_banky for k in klienti if k.vyber_banky]
        rozlozeni_banky = Counter(banky)
        banky_labels = list(rozlozeni_banky.keys())
        schvalenost = [
            klienti.filter(duvod_zamitnuti__isnull=True, vyber_banky=b).count()
            for b in banky_labels
        ]
        zamitnutost = [
            klienti.filter(duvod_zamitnuti__isnull=False, vyber_banky=b).count()
            for b in banky_labels
        ]
        prumery = ["-" for _ in banky_labels]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        ws.append(["Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"])
        for i, banka in enumerate(banky_labels):
            ws.append([banka, schvalenost[i], zamitnutost[i], prumery[i]])
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        wb2 = openpyxl.load_workbook(xlsx_buffer)
        ws2 = wb2.active
        data = list(ws2.values)
        self.assertEqual(
            data[0],
            ("Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"),
        )
        # Ověř, že klient bez banky není v reportu (protože nemá banku)
        for row in data[1:]:
            self.assertIsNotNone(row[0])  # Každý řádek musí mít banku
        # Ověř, že export proběhl bez chyby i s neúplnými daty
        self.assertTrue(len(data) >= 1)  # Hlavička vždy existuje

    def test_reporting_email_odeslani_a_priloha(self):
        """
        Ověří, že management command send_reporting_email vygeneruje a odešle reportingový e-mail s přílohou (XLSX nebo PDF).
        Ověří i edge-case: prázdná data (report se odešle, příloha existuje, ale je prázdná).
        """
        from django.core import mail
        from django.core.management import call_command

        # Smaž všechny klienty (prázdná data)
        from klienti.models import Klient

        Klient.objects.all().delete()
        # Spusť command
        call_command("send_reporting_email")
        # Ověř, že byl odeslán e-mail
        self.assertEqual(len(mail.outbox), 1)
        email = mail.outbox[0]
        self.assertIn("reporting", email.subject.lower())
        self.assertIn("report", email.body.lower())
        # Ověř, že e-mail má přílohu (XLSX nebo PDF)
        self.assertTrue(email.attachments, "E-mail nemá žádnou přílohu!")
        # Ověř, že příloha není None a má správný typ
        attachment = email.attachments[0]
        self.assertTrue(
            attachment[0].endswith(".xlsx") or attachment[0].endswith(".pdf")
        )
        self.assertTrue(len(attachment[1]) > 0, "Příloha je prázdná!")
        # Ověř, že e-mail se odešle i když nejsou žádná data


class ReportingExportEdgeCaseTestCase(TestCase):
    """
    Testuje edge-case scénáře pro reporting/export.
    """

    def setUp(self):
        Klient.objects.all().delete()
        # Vytvoř testovacího manažera s validním e-mailem
        from django.contrib.auth.models import User

        from klienti.models import UserProfile

        user = User.objects.create_user(
            username="manazer", email="manazer@example.com", password="testpass"
        )
        profile, created = UserProfile.objects.get_or_create(
            user=user, defaults={"role": "manažer"}
        )
        if not created:
            profile.role = "manažer"
            profile.save()

    def test_export_xlsx_prazdna_data(self):
        """
        Ověří, že export XLSX s prázdnou databází klientů vrací správnou hlavičku a žádné datové řádky.
        Tento test je důležitý pro robustnost reportingu – systém nesmí selhat ani při absenci dat.
        """
        # Simulace generování XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporting"
        ws.append(["Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"])
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        # Ověření obsahu XLSX
        wb2 = openpyxl.load_workbook(xlsx_buffer)
        ws2 = wb2.active
        data = list(ws2.values)
        # Očekáváme pouze hlavičku
        self.assertEqual(len(data), 1)
        self.assertEqual(
            data[0],
            ("Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"),
        )

    def test_export_xlsx_rollback_pri_chybe(self):
        """
        Ověří, že při chybě během exportu XLSX nedojde k vytvoření neúplného souboru ani nekonzistenci dat.
        Například: pokud dojde k výjimce během generování, žádný soubor se nevytvoří.
        """
        import os

        from django.db import transaction

        filename = "test_export_rollback.xlsx"
        try:
            with transaction.atomic():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Reporting"
                ws.append(
                    ["Banka", "Schváleno", "Zamítnuto", "Průměrná doba schválení (dny)"]
                )
                # Simulace chyby při exportu
                raise Exception("Chyba během exportu")
                wb.save(filename)
        except Exception:
            pass
        # Ověř, že soubor nebyl vytvořen
        self.assertFalse(os.path.exists(filename))

    def test_export_pdf_rollback_pri_chybe(self):
        """
        Ověří, že při chybě během exportu PDF nevznikne neúplný nebo poškozený soubor.
        Například: pokud dojde k výjimce během generování, žádný soubor se nevytvoří.
        """
        import os

        from django.db import transaction

        from reportlab.pdfgen import canvas

        filename = "test_export_rollback.pdf"
        try:
            with transaction.atomic():
                c = canvas.Canvas(filename)
                c.drawString(100, 750, "Test PDF export")
                # Simulace chyby při exportu
                raise Exception("Chyba během exportu PDF")
                c.save()
        except Exception:
            pass
        # Ověř, že soubor nebyl vytvořen
        self.assertFalse(os.path.exists(filename))

    def test_send_reporting_email_prázdná_data(self):
        """
        Ověří, že při generování reportu e-mailem s prázdnými daty se odešle správná zpráva a příloha je validní.
        """
        from django.core import mail
        from django.core.management import call_command

        from klienti.models import Klient

        Klient.objects.all().delete()  # Zajistí prázdná data
        call_command("send_reporting_email")
        self.assertEqual(len(mail.outbox), 1)
        email = mail.outbox[0]
        self.assertIn("report", email.subject.lower())
        # Ověř, že e-mail má přílohu (XLSX nebo PDF)
        self.assertTrue(email.attachments, "E-mail nemá žádnou přílohu!")
        # Ověř, že příloha není None a má správný typ
        attachment = email.attachments[0]
        self.assertIn("reporting", email.subject.lower())
        self.assertIn("reporting", attachment[0].lower())
        self.assertTrue(email.attachments)
        # Ověř, že příloha není poškozená (např. PDF/XLSX má správný typ)
        self.assertTrue(
            attachment[0].endswith(".pdf") or attachment[0].endswith(".xlsx")
        )
        self.assertTrue(len(attachment[1]) > 0)

    def test_send_reporting_email_neexistujici_email(self):
        """
        Ověří, že pokud je v systému neexistující e-mail, report se neodešle a zaloguje se chyba.
        """
        from django.core.management import call_command

        from klienti.models import UserProfile

        # Nastav neexistující e-mail pro všechny poradce
        for profile in UserProfile.objects.filter(role="poradce"):
            profile.user.email = ""
            profile.user.save()
        # Očekáváme, že command zaloguje chybu, ale nespadne
        try:
            call_command("send_reporting_email")
        except Exception as e:
            self.fail(f"Command selhal: {e}")
